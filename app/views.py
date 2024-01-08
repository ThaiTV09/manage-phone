from django.shortcuts import render,redirect
from django.http import HttpResponse,JsonResponse
from .models import *
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from app.forms import PhoneForm, PhoneFormUpdate
import csv
from .models import Phone
from django.db.models import Avg, Sum, Min, Max, Count
from django.utils import timezone

#Create, edit, delete
def addnew(request):  
    if request.method == "POST":  
        form = PhoneForm(request.POST)  
        if form.is_valid():  
            try:  
                form.save()  
                return redirect('/')  
            except:  
                pass
    else:  
        form = PhoneForm()  
    return render(request,'app/index.html',{'form':form}) 
 
def index(request):  
    phones = Phone.objects.all()  
    return render(request,"app/show.html",{'phones':phones})

def warning_spider(request):  
    phones = Phone.objects.filter(system_id_id='1', money__gt=10000).order_by('date_connect')
    return render(request,"app/warning_spider.html",{'phones':phones})
def warning_dspm(request):  
    phones = Phone.objects.filter(system_id_id='3', money__gt=10000).order_by('date_connect')
    return render(request,"app/warning_dspm.html",{'phones':phones})
def warning_scada(request):  
    phones = Phone.objects.filter(system_id_id='2', money__lt=30000).order_by('date_connect')
    return render(request,"app/warning_scada.html",{'phones':phones})
def warning_mtmn(request):  
    phones = Phone.objects.filter(system_id_id='4', money__gt=10000).order_by('date_connect')
    return render(request,"app/warning_mtmn.html",{'phones':phones})
def warning_data(request):  
    phones = Phone.objects.filter(system_id_id='5', money__gt=10000).order_by('date_connect')
    return render(request,"app/warning_data.html",{'phones':phones})    

def service(request):    
    phoneitems = PhoneItem.objects.filter(data__lte=0)
    if request.method == "POST":
        fromdate=request.POST.get('fromdate')
        todate=request.POST.get('todate')
        phoneitems = PhoneItem.objects.filter(date_use__range=[fromdate,todate],data__lte=0)
    return render(request,"app/service.html",{'phoneitems':phoneitems})
def service_low(request):    
    phoneitems = PhoneItem.objects.filter(data__lt=10000, data__gt=1) 
    if request.method == "POST":
        searched=request.POST.get('searched')
        fromdate=request.POST.get('fromdate')
        todate=request.POST.get('todate')
        phoneitems = PhoneItem.objects.filter(date_use__range=[fromdate,todate],data__lte=searched,data__gt=0)
    return render(request,"app/service_low.html",{'phoneitems':phoneitems})    
def service_high(request):    
    phoneitems = PhoneItem.objects.filter(data__gt=500000)  
    if request.method == "POST":
        searched=request.POST.get('searched')
        fromdate=request.POST.get('fromdate')
        todate=request.POST.get('todate')
        phoneitems = PhoneItem.objects.filter(date_use__range=[fromdate,todate],data__gt=searched)
    return render(request,"app/service_high.html",{'phoneitems':phoneitems})    
    
def edit(request, serial):  
    phone = Phone.objects.get(serial=serial)
    form = PhoneFormUpdate()      
    departments = Department.objects.filter()  
    return render(request,'app/edit.html', {'form': form, 'phone':phone, 'departments': departments})  
 
def update(request, serial):
    if request.method == 'POST':  
        phone = Phone.objects.get(serial=serial)
        departments = Department.objects.filter()  
        form = PhoneFormUpdate(request.POST, instance = phone)  
        if form.is_valid():  
            form.save()  
            return redirect("/")
    form = PhoneFormUpdate()  
    return render(request, 'app/edit.html', {'form': form, 'phone': phone, 'departments': departments})  
     
def destroy(request, serial):  
    phone = Phone.objects.get(serial=serial)  
    phone.delete()  
    return redirect("/")  
# Create your views here.


def home(request):
    phone_list = Phone.objects.filter().order_by('serial')
    departments = Department.objects.filter()
    context={'phone_list':phone_list, 'departments':departments}
    return render(request,'app/home.html',context)

def base(request):
    context={}
    return render(request,'app/base.html',context)

def export_to_excel(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="phones.xlsx"'
    wb = Workbook()
    ws = wb.active
    ws.title = "Phone"
    # Add headers
    headers = ["SĐT", "Serial", "Gói cước","đơn giá","Ngày hòa mạng","Đơn vị sử dụng"]
    ws.append(headers)  
    # Add data from the model
    phones = Phone.objects.all()
    for id in phones:
        ws.append([str(id.phone), str(id.serial), id.pack, id.money, id.date_connect, id.department_id.department])
    #last_cell = 100
    #for col in range(1,2):
    #    for row in range(1, last_cell):
    #        ws.cell(column=col, row=row).number_format = "@"  # Changing format to TEXT
    wb.save(response)
    return response 

def export_query_to_csv(request):
    data = Phone.objects.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="phone.csv"'

    writer = csv.writer(response)
    writer.writerow(["SĐT", "Serial", "Gói cước","đơn giá","Ngày hòa mạng","Đơn vị sử dụng"])  # CSV header

    for id in data:
        writer.writerow([str(id.phone), str(id.serial), id.pack, id.money, id.date_connect, id.department_id.department])

    return response


#Dashboard View
def chart_system(request):
    totalserrial = Phone.objects.filter().count()
    totalserrial = int(totalserrial)
    print('Number of Thuê bao ',totalserrial)
    totalmobifone = Phone.objects.filter(network_id_id='1').count()
    totalmobifone = int(totalmobifone)
    print('Number of MOBIFONE ',totalmobifone)
    totalviettel = Phone.objects.filter(network_id_id='2').count()
    totalviettel = int(totalviettel)
    print('Number of VIETTEL ',totalviettel)
    totalvinaphone = Phone.objects.filter(network_id_id='3').count()
    totalvinaphone = int(totalvinaphone)
    print('Number of VINAPHONE ',totalvinaphone)
#Số lượng thuê bao theo chương trình      
    spider = Phone.objects.filter(system_id_id='1').count()
    spider = int(spider)
    print('Number of RF-Spider in System',spider)
    scada = Phone.objects.filter(system_id_id='2').count()
    scada = int(scada)
    print('Number of SCADA in System',scada)
    dspm = Phone.objects.filter(system_id_id='3').count()
    dspm = int(dspm)
    print('Number of DSPM in System',dspm)
    srfi = Phone.objects.filter(system_id_id='4').count()
    srfi = int(srfi)
    print('Number of SRFI in System',srfi)
    mtmn = Phone.objects.filter(system_id_id='5').count()
    mtmn = int(mtmn)
    print('Number of Điện MTMN in System',mtmn)
    home_phone = Phone.objects.filter(system_id_id='6').count()
    home_phone = int(home_phone)
    print('Number of Home_phone in Department',home_phone)
    no_system = Phone.objects.filter(system_id_id='7').count()
    no_system = int(no_system)
    print('Number of No System in Department',no_system)
#Số lượng thuê bao theo đơn vị
    qtpc = Phone.objects.filter(department_id_id='1').count()
    qtpc = int(qtpc)
    print('Number of QTPC in Department',qtpc)
    pc02aa = Phone.objects.filter(department_id_id='2').count()
    pc02aa = int(pc02aa)
    print('Number of Đông Hà in Department',pc02aa)
    pc02bb = Phone.objects.filter(department_id_id='3').count()
    pc02bb = int(pc02bb)
    print('Number of Thành Cổ in Department',pc02bb)
    pc02cc = Phone.objects.filter(department_id_id='4').count()
    pc02cc = int(pc02cc)
    print('Number of Gio Linh in Department',pc02cc)
    pc02dd = Phone.objects.filter(department_id_id='5').count()
    pc02dd = int(pc02dd)
    print('Number of Vĩnh Linh in Department',pc02dd)
    pc02ff = Phone.objects.filter(department_id_id='6').count()
    pc02ff = int(pc02ff)
    print('Number of Khe Sanh in Department',pc02ff)
    pc02gg = Phone.objects.filter(department_id_id='7').count()
    pc02gg = int(pc02gg)
    print('Number of Cam lộ in Department',pc02gg)
    pc02kk = Phone.objects.filter(department_id_id='8').count()
    pc02kk = int(pc02kk)
    print('Number of Hải Lăng in Department',pc02kk)
    pc02hh = Phone.objects.filter(department_id_id='9').count()
    pc02hh = int(pc02hh)
    print('Number of Triệu Phong in Department',pc02hh)
    pc02ll = Phone.objects.filter(department_id_id='10').count()
    pc02ll = int(pc02ll)
    print('Number of Đakrông in Department',pc02ll)
    pc02mm = Phone.objects.filter(department_id_id='11').count()
    pc02mm = int(pc02mm)
    print('Number of TĐCC in Department',pc02mm)
    qlvh = Phone.objects.filter(department_id_id='14').count()
    qlvh = int(qlvh)
    print('Number of Cồn Cỏ in Department',qlvh)

    depart_list = ['Công ty', 'ĐL Đông Hà', 'ĐL Thành Cổ', 'ĐL Gio Linh', 'ĐL Vĩnh Linh', 'ĐL Khe Sanh', 'ĐL Cam Lộ', 'ĐL Hải Lăng', 'ĐL Triệu Phong', 'ĐL Đakrông', 'TĐ Cồn Cỏ', 'Đội QLVH']
    numderpart_list = [qtpc, pc02aa, pc02bb, pc02cc, pc02dd, pc02ff, pc02gg, pc02kk, pc02hh, pc02ll, pc02mm, qlvh]
    course_list = ['RF-Spider', 'SCADA', 'DSPM', 'SRFI', 'MTMN', 'Phone', 'Chưa phân bổ']
    number_list = [spider, scada, dspm, srfi, mtmn, home_phone, no_system]

    context = {'totalmobifone':totalmobifone,'totalviettel':totalviettel,'totalvinaphone':totalvinaphone, 'totalserrial':totalserrial, 'course_list':course_list, 'number_list':number_list, 'depart_list':depart_list, 'numderpart_list':numderpart_list}
    return render(request,"app/chart_system.html", context)

def chart_fee(request):
    
    totalfee = Phone.objects.all().aggregate(Sum("money"))['money__sum'] or 0
    totalfee = int(totalfee)
    totalfee = format(totalfee,',') 
    totalmobifone = Phone.objects.filter(network_id_id='1').aggregate(Sum("money"))['money__sum'] or 0
    totalmobifone = int(totalmobifone)
    totalmobifone = format(totalmobifone,',') 
    totalviettel = Phone.objects.filter(network_id_id='2').aggregate(Sum("money"))['money__sum'] or 0
    totalviettel = int(totalviettel)
    totalviettel = format(totalviettel,',') 
    totalvinaphone = Phone.objects.filter(network_id_id='3').aggregate(Sum("money"))['money__sum'] or 0
    totalvinaphone = int(totalvinaphone)
    totalvinaphone = format(totalvinaphone,',') 
#Chi phí thuê bao theo chương trình      
    spider = Phone.objects.filter(system_id_id='1').aggregate(Sum("money"))['money__sum']
    print('Number of RF-Spider in System',spider)
    scada = Phone.objects.filter(system_id_id='2').aggregate(Sum("money"))['money__sum']
    print('Number of SCADA in System',scada)
    dspm = Phone.objects.filter(system_id_id='3').aggregate(Sum("money"))['money__sum']
    print('Number of DSPM in System',dspm)
    srfi = Phone.objects.filter(system_id_id='4').aggregate(Sum("money"))['money__sum']
    print('Number of SRFI in System',srfi)
    mtmn = Phone.objects.filter(system_id_id='5').aggregate(Sum("money"))['money__sum']
    print('Number of Điện MTMN in System',mtmn)
    home_phone = Phone.objects.filter(system_id_id='6').aggregate(Sum("money"))['money__sum']
    print('Number of Home_phone in Department',home_phone)
    no_system = Phone.objects.filter(system_id_id='7').aggregate(Sum("money"))['money__sum']
    print('Number of No System in Department',no_system)
#Số lượng thuê bao theo đơn vị
    qtpc = Phone.objects.filter(department_id_id='1').aggregate(Sum("money"))['money__sum']
    print('Number of QTPC in Department',qtpc)
    pc02aa = Phone.objects.filter(department_id_id='2').aggregate(Sum("money"))['money__sum']
    print('Number of Đông Hà in Department',pc02aa)
    pc02bb = Phone.objects.filter(department_id_id='3').aggregate(Sum("money"))['money__sum']
    print('Number of Thành Cổ in Department',pc02bb)
    pc02cc = Phone.objects.filter(department_id_id='4').aggregate(Sum("money"))['money__sum']
    print('Number of Gio Linh in Department',pc02cc)
    pc02dd = Phone.objects.filter(department_id_id='5').aggregate(Sum("money"))['money__sum']
    print('Number of Vĩnh Linh in Department',pc02dd)
    pc02ff = Phone.objects.filter(department_id_id='6').aggregate(Sum("money"))['money__sum']
    print('Number of Khe Sanh in Department',pc02ff)
    pc02gg = Phone.objects.filter(department_id_id='7').aggregate(Sum("money"))['money__sum']
    print('Number of Cam lộ in Department',pc02gg)
    pc02kk = Phone.objects.filter(department_id_id='8').aggregate(Sum("money"))['money__sum']
    print('Number of Hải Lăng in Department',pc02kk)
    pc02hh = Phone.objects.filter(department_id_id='9').aggregate(Sum("money"))['money__sum']
    print('Number of Triệu Phong in Department',pc02hh)
    pc02ll = Phone.objects.filter(department_id_id='10').aggregate(Sum("money"))['money__sum']
    print('Number of Đakrông in Department',pc02ll)
    pc02mm = Phone.objects.filter(department_id_id='11').aggregate(Sum("money"))['money__sum']
    print('Number of TĐCC in Department',pc02mm)
    qlvh = Phone.objects.filter(department_id_id='14').aggregate(Sum("money"))['money__sum']
    print('Number of Cồn Cỏ in Department',qlvh)

    depart_list = ['Công ty', 'ĐL Đông Hà', 'ĐL Thành Cổ', 'ĐL Gio Linh', 'ĐL Vĩnh Linh', 'ĐL Khe Sanh', 'ĐL Cam Lộ', 'ĐL Hải Lăng', 'ĐL Triệu Phong', 'ĐL Đakrông', 'TĐ Cồn Cỏ', 'Đội QLVH']
    numderpart_list = [qtpc, pc02aa, pc02bb, pc02cc, pc02dd, pc02ff, pc02gg, pc02kk, pc02hh, pc02ll, pc02mm, qlvh]
    course_list = ['RF-Spider', 'SCADA', 'DSPM', 'SRFI', 'MTMN', 'Phone', 'Chưa phân bổ']
    number_list = [spider, scada, dspm, srfi, mtmn, home_phone, no_system]

    context = {'totalmobifone':totalmobifone,'totalviettel':totalviettel,'totalvinaphone':totalvinaphone, 'totalfee':totalfee, 'course_list':course_list, 'number_list':number_list, 'depart_list':depart_list, 'numderpart_list':numderpart_list}
    return render(request,"app/chart_fee.html", context)

def chart_fee_system(request):
    departments = Department.objects.filter()
    year = timezone.now().year - 1
    active_department = request.GET.get('chat_fee_system','')
    if active_department:
        depart = Department.objects.filter(department_id=active_department).values()
        totalserial = Phone.objects.filter(department_id_id = active_department).count()
        totalmobi = Phone.objects.filter(department_id_id = active_department, network_id_id='1').count()
        totalvtel = Phone.objects.filter(department_id_id = active_department, network_id_id='2').count()
        totalvina = Phone.objects.filter(department_id_id = active_department, network_id_id='3').count()
        fee_totalserial = Phone.objects.filter(department_id_id = active_department).aggregate(Sum("money"))['money__sum'] or 0
        fee_totalserial = int(fee_totalserial)
        fee_totalserial = format(fee_totalserial,',')
        fee_totalmobi = Phone.objects.filter(department_id_id = active_department, network_id_id='1').aggregate(Sum("money"))['money__sum'] or 0
        fee_totalmobi = int(fee_totalmobi)
        fee_totalmobi = format(fee_totalmobi,',')        
        fee_totalvtel = Phone.objects.filter(department_id_id = active_department, network_id_id='2').aggregate(Sum("money"))['money__sum'] or 0
        fee_totalvtel = int(fee_totalvtel)
        fee_totalvtel = format(fee_totalvtel,',') 
        fee_totalvina = Phone.objects.filter(department_id_id = active_department, network_id_id='3').aggregate(Sum("money"))['money__sum'] or 0
        fee_totalvina = int(fee_totalvina)
        fee_totalvina = format(fee_totalvina,',') 
        n_spider = Phone.objects.filter(department_id_id = active_department, system_id_id='1').count()
        n_scada = Phone.objects.filter(department_id_id = active_department, system_id_id='2').count()
        n_dspm = Phone.objects.filter(department_id_id = active_department,system_id_id='3').count()
        n_srfi = Phone.objects.filter(department_id_id = active_department,system_id_id='4').count()
        n_nlmt = Phone.objects.filter(department_id_id = active_department,system_id_id='5').count()
        n_home_phone = Phone.objects.filter(department_id_id = active_department,system_id_id='6').count()
        spider = Phone.objects.filter(department_id_id = active_department, system_id_id='1').aggregate(Sum("money"))['money__sum'] or 0
        scada = Phone.objects.filter(department_id_id = active_department, system_id_id='2').aggregate(Sum("money"))['money__sum'] or 0
        dspm = Phone.objects.filter(department_id_id = active_department,system_id_id='3').aggregate(Sum("money"))['money__sum'] or 0
        srfi = Phone.objects.filter(department_id_id = active_department,system_id_id='4').aggregate(Sum("money"))['money__sum'] or 0
        nlmt = Phone.objects.filter(department_id_id = active_department,system_id_id='5').aggregate(Sum("money"))['money__sum'] or 0
        home_phone = Phone.objects.filter(department_id_id = active_department,system_id_id='6').aggregate(Sum("money"))['money__sum'] or 0
        count=Phone.objects.filter(department_id_id = active_department,date_connect__year=year).count()
        money=Phone.objects.filter(department_id_id = active_department,date_connect__year=year).aggregate(Sum("money"))['money__sum'] or 0
        money = int(money)
        money = format(money,',') 

        course_list = ['RF-Spider', 'SCADA', 'DSPM', 'SRFI', 'MTMN', 'Phone']
        number_list = [spider, scada, dspm, srfi, nlmt, home_phone]
        n_course_list = ['RF-Spider', 'SCADA', 'DSPM', 'SRFI', 'MTMN', 'Phone']
        n_number_list = [n_spider, n_scada, n_dspm, n_srfi, n_nlmt, n_home_phone]

    context={'money':money,'count':count,'course_list':course_list,'number_list':number_list,'n_course_list':n_course_list,'n_number_list':n_number_list,'depart':depart,'fee_totalvina':fee_totalvina,'fee_totalvtel':fee_totalvtel,'fee_totalmobi':fee_totalmobi,'fee_totalserial':fee_totalserial,'totalvina':totalvina,'totalmobi':totalmobi,'totalvtel':totalvtel, 'departments':departments, 'active_department':active_department, 'totalserial':totalserial}
    return render(request,'app/chart_fee_system.html',context)